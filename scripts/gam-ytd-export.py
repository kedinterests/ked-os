#!/usr/bin/env python3
"""Export YTD advertiser performance to CSV and Excel."""

import json, csv, os

with open('secrets/ytd-report-raw.json') as f:
    data = json.load(f)

rows = data.get('rows', [])

MONTH_MAP = {
    '1512': '2026-01 Jan', '1513': '2026-02 Feb', '1514': '2026-03 Mar',
    '1515': '2026-04 Apr', '1516': '2026-05 May', '1517': '2026-06 Jun',
}

parsed = []
for row in rows:
    dv = row['dimensionValues']
    mv = row['metricValueGroups'][0]['primaryValues']
    month_raw = str(dv[0].get('intValue', ''))
    advertiser = dv[1].get('stringValue', '')
    order = dv[2].get('stringValue', '')
    imp = int(mv[0].get('intValue', 0))
    clicks = int(mv[1].get('intValue', 0))
    ctr = round(mv[2].get('doubleValue', 0) * 100, 4)
    parsed.append({
        'Month': MONTH_MAP.get(month_raw, month_raw),
        'Advertiser': advertiser,
        'Order': order,
        'Impressions': imp,
        'Clicks': clicks,
        'CTR %': ctr,
    })

parsed.sort(key=lambda r: (r['Advertiser'], r['Month']))

# CSV
csv_path = 'KED-YTD-Ad-Performance-2026.csv'
with open(csv_path, 'w', newline='') as f:
    writer = csv.DictWriter(f, fieldnames=['Month','Advertiser','Order','Impressions','Clicks','CTR %'])
    writer.writeheader()
    writer.writerows(parsed)
print(f'CSV saved: {os.path.abspath(csv_path)}')

# Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'YTD Performance'

    NAVY = '0D1B2A'
    TEAL = '00B4D8'
    LIGHT = 'F8F9FA'
    ALT   = 'EEF6F9'

    headers = ['Month', 'Advertiser', 'Order', 'Impressions', 'Clicks', 'CTR %']
    col_widths = [18, 28, 30, 16, 10, 10]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill('solid', fgColor=NAVY)
        cell.alignment = Alignment(horizontal='center' if col > 1 else 'left', vertical='center')
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    thin = Side(style='thin', color='DEE2E6')
    border = Border(bottom=Side(style='thin', color='DEE2E6'))

    for i, row in enumerate(parsed, 2):
        fill_color = LIGHT if i % 2 == 0 else ALT
        for col, key in enumerate(headers, 1):
            val = row[key]
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill = PatternFill('solid', fgColor=fill_color)
            cell.border = border
            if col == 1:
                cell.alignment = Alignment(horizontal='left')
            elif col in (4, 5):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            elif col == 6:
                cell.number_format = '0.000"%"'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')

    # Totals row
    total_row = len(parsed) + 2
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=total_row, column=1).fill = PatternFill('solid', fgColor='D0EEF5')

    total_imp = sum(r['Impressions'] for r in parsed)
    total_clicks = sum(r['Clicks'] for r in parsed)
    total_ctr = round(total_clicks / total_imp * 100, 4) if total_imp else 0

    for col, val in [(4, total_imp), (5, total_clicks), (6, total_ctr)]:
        cell = ws.cell(row=total_row, column=col, value=val)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='D0EEF5')
        cell.alignment = Alignment(horizontal='right')
        cell.number_format = '#,##0' if col < 6 else '0.000"%"'

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:F{len(parsed)+1}'

    xl_path = 'KED-YTD-Ad-Performance-2026.xlsx'
    wb.save(xl_path)
    print(f'Excel saved: {os.path.abspath(xl_path)}')

except ImportError:
    import subprocess
    subprocess.run(['pip3', 'install', '--break-system-packages', 'openpyxl'], check=True)
    print('openpyxl installed — run the script again.')
