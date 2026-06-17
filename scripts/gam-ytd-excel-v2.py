#!/usr/bin/env python3
"""YTD Excel — one row per advertiser, months merged across top."""

import json, os
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

with open('secrets/ytd-report-raw.json') as f:
    data = json.load(f)

MONTH_MAP = {
    '1512': 'Jan', '1513': 'Feb', '1514': 'Mar',
    '1515': 'Apr', '1516': 'May', '1517': 'Jun',
}
MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun']

by_adv = defaultdict(lambda: defaultdict(lambda: [0, 0]))
for row in data.get('rows', []):
    dv = row['dimensionValues']
    mv = row['metricValueGroups'][0]['primaryValues']
    month = MONTH_MAP.get(str(dv[0].get('intValue', '')), '')
    adv = dv[1].get('stringValue', '')
    imp = int(mv[0].get('intValue', 0))
    clicks = int(mv[1].get('intValue', 0))
    by_adv[adv][month][0] += imp
    by_adv[adv][month][1] += clicks

ADVERTISERS = sorted(by_adv.keys())

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'YTD Performance'

NAVY   = '0D1B2A'
TEAL   = '00B4D8'
TEAL_L = 'D0EEF5'
WHITE  = 'FFFFFF'
OFF    = 'F8F9FA'
ALT    = 'EEF6F9'
GRAY   = 'DEE2E6'
MID    = '6C757D'
TOTAL_BG = 'C8E6EF'

def fill(hex): return PatternFill('solid', fgColor=hex)
def font(bold=False, color='000000', size=10): return Font(bold=bold, color=color, size=size)
def align(h='left', v='center', wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
thin = Side(style='thin', color=GRAY)
med  = Side(style='medium', color=TEAL)

def border(left=None, right=None, top=None, bottom=None):
    return Border(left=left, right=right, top=top, bottom=bottom)

# Column layout:
# Col 1: Advertiser
# Then 3 cols per month (Impressions, Clicks, CTR%) x 6 months
# Then 3 cols for Total
# = 1 + 6*3 + 3 = 22 columns

ADV_COL = 1
MONTH_STARTS = {m: ADV_COL + 1 + i*3 for i, m in enumerate(MONTHS)}
TOTAL_START  = ADV_COL + 1 + 6*3

# Row 1: "Advertiser" header + merged month headers
ws.row_dimensions[1].height = 20
ws.row_dimensions[2].height = 16
ws.row_dimensions[3].height = 14

# Advertiser header — spans rows 1-2
ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
c = ws.cell(row=1, column=1, value='Advertiser')
c.font = font(bold=True, color=WHITE, size=10)
c.fill = fill(NAVY)
c.alignment = align('left', 'center')
ws.column_dimensions['A'].width = 26

# Month merged headers
for m in MONTHS:
    sc = MONTH_STARTS[m]
    label = m if m != 'Jun' else 'Jun *'
    ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=sc+2)
    c = ws.cell(row=1, column=sc, value=label)
    c.font = font(bold=True, color=WHITE, size=10)
    c.fill = fill(TEAL if m != 'Jun' else '0090B0')
    c.alignment = align('center', 'center')

# Total merged header
ws.merge_cells(start_row=1, start_column=TOTAL_START, end_row=1, end_column=TOTAL_START+2)
c = ws.cell(row=1, column=TOTAL_START, value='Total')
c.font = font(bold=True, color=NAVY, size=10)
c.fill = fill(TEAL_L)
c.alignment = align('center', 'center')

# Sub-headers row 2
sub_labels = ['Impressions', 'Clicks', 'CTR %']
sub_widths  = [14, 8, 8]
for m in MONTHS:
    sc = MONTH_STARTS[m]
    for i, (lbl, w) in enumerate(zip(sub_labels, sub_widths)):
        c = ws.cell(row=2, column=sc+i, value=lbl)
        c.font = font(bold=True, color=WHITE, size=8)
        c.fill = fill(NAVY)
        c.alignment = align('right', 'center')
        ws.column_dimensions[get_column_letter(sc+i)].width = w

for i, (lbl, w) in enumerate(zip(sub_labels, sub_widths)):
    c = ws.cell(row=2, column=TOTAL_START+i, value=lbl)
    c.font = font(bold=True, color=NAVY, size=8)
    c.fill = fill(TEAL_L)
    c.alignment = align('right', 'center')
    ws.column_dimensions[get_column_letter(TOTAL_START+i)].width = w

# Data rows
totals_by_month = defaultdict(lambda: [0, 0])
grand = [0, 0]

for row_idx, adv in enumerate(ADVERTISERS, 3):
    ws.row_dimensions[row_idx].height = 15
    bg = OFF if row_idx % 2 == 1 else ALT

    c = ws.cell(row=row_idx, column=1, value=adv)
    c.font = font(bold=False, size=9)
    c.fill = fill(bg)
    c.alignment = align('left', 'center')

    adv_imp = adv_clicks = 0
    for m in MONTHS:
        imp, clicks = by_adv[adv][m]
        ctr = round(clicks / imp * 100, 4) if imp else 0
        adv_imp += imp; adv_clicks += clicks
        totals_by_month[m][0] += imp; totals_by_month[m][1] += clicks
        sc = MONTH_STARTS[m]

        ci = ws.cell(row=row_idx, column=sc, value=imp if imp else None)
        ci.font = font(size=9); ci.fill = fill(bg)
        ci.number_format = '#,##0'; ci.alignment = align('right', 'center')

        cc = ws.cell(row=row_idx, column=sc+1, value=clicks if imp else None)
        cc.font = font(size=9); cc.fill = fill(bg)
        cc.number_format = '#,##0'; cc.alignment = align('right', 'center')

        ct = ws.cell(row=row_idx, column=sc+2, value=ctr if imp else None)
        ct.font = font(size=9); ct.fill = fill(bg)
        ct.number_format = '0.000"%"'; ct.alignment = align('right', 'center')

    grand[0] += adv_imp; grand[1] += adv_clicks
    total_ctr = round(adv_clicks / adv_imp * 100, 4) if adv_imp else 0

    ti = ws.cell(row=row_idx, column=TOTAL_START, value=adv_imp)
    ti.font = font(bold=True, size=9); ti.fill = fill(TEAL_L)
    ti.number_format = '#,##0'; ti.alignment = align('right', 'center')

    tc = ws.cell(row=row_idx, column=TOTAL_START+1, value=adv_clicks)
    tc.font = font(bold=True, size=9); tc.fill = fill(TEAL_L)
    tc.number_format = '#,##0'; tc.alignment = align('right', 'center')

    tt = ws.cell(row=row_idx, column=TOTAL_START+2, value=total_ctr)
    tt.font = font(bold=True, size=9); tt.fill = fill(TEAL_L)
    tt.number_format = '0.000"%"'; tt.alignment = align('right', 'center')

# Totals row
total_row = len(ADVERTISERS) + 3
ws.row_dimensions[total_row].height = 16

c = ws.cell(row=total_row, column=1, value='All advertisers')
c.font = font(bold=True, size=9, color=NAVY); c.fill = fill(TOTAL_BG)
c.alignment = align('left', 'center')

for m in MONTHS:
    imp, clicks = totals_by_month[m]
    ctr = round(clicks / imp * 100, 4) if imp else 0
    sc = MONTH_STARTS[m]
    for col, val, fmt in [(sc, imp, '#,##0'), (sc+1, clicks, '#,##0'), (sc+2, ctr, '0.000"%"')]:
        cell = ws.cell(row=total_row, column=col, value=val)
        cell.font = font(bold=True, size=9, color=NAVY)
        cell.fill = fill(TOTAL_BG)
        cell.number_format = fmt
        cell.alignment = align('right', 'center')

grand_ctr = round(grand[1] / grand[0] * 100, 4) if grand[0] else 0
for col, val, fmt in [(TOTAL_START, grand[0], '#,##0'), (TOTAL_START+1, grand[1], '#,##0'), (TOTAL_START+2, grand_ctr, '0.000"%"')]:
    cell = ws.cell(row=total_row, column=col, value=val)
    cell.font = font(bold=True, size=9, color=NAVY)
    cell.fill = fill(TOTAL_BG)
    cell.number_format = fmt
    cell.alignment = align('right', 'center')

# Freeze header rows
ws.freeze_panes = 'B3'

# Note row
note_row = total_row + 2
ws.cell(row=note_row, column=1, value='* Jun = partial month (through Jun 17, 2026)').font = Font(size=8, color=MID, italic=True)

OUT = 'KED-YTD-Ad-Performance-2026.xlsx'
wb.save(OUT)
print(f'Saved: {os.path.abspath(OUT)}')
